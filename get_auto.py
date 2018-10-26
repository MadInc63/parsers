import tqdm
import time
import user_agent
import sqlite3
import requests
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from queue import Queue
from lxml import html
from openpyxl import Workbook, load_workbook
from threading import Thread


def get_request(url):
    desktop_agent = user_agent.generate_user_agent(os=('win', 'mac', 'linux'))
    user_agent_header = {'User-Agent': desktop_agent}
    header = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/67.0.3396.99 Mobile Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,'
                  'application/xml;q=0.9,image/webp,'
                  'image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'ru, en; q =0.9',
        'Connection': 'keep-alive'
    }
    header.update(user_agent_header)
    response = requests.get(url, headers=header, timeout=30)
    if response.status_code == requests.codes.ok:
        return response
    else:
        return None


def get_page_count(url):
    driver.get(url)
    wait.until(ec.element_to_be_clickable((
        By.CSS_SELECTOR, '#searchPagination > div > nav'
    )))
    elements = driver.find_elements_by_xpath(
        '//nav[@class="unstyle pager"]/span[@class="page-item mhide"]/a'
    )
    elements = [
        int(element.text.replace(' ', '')) for element
        in elements if element.text
    ]
    return max(elements)


def get_ads_url(url):
    response = get_request(url)
    tree = html.fromstring(response.content)
    url_list = tree.xpath(
        '//div[@id="searchResults"]//a[@class="address"]/@href'
    )
    return url_list


def get_ads_info(url):
    response = get_request(url)
    tree = html.fromstring(response.content)
    ads_title = tree.xpath('//div[@class="heading"]/h1//text()')
    ads_title = [text.strip() for text in ads_title if text != ' ']
    ads_title = ' '.join(ads_title)
    check_offers = tree.xpath('//div[@class="item-phone"]/a/span/text()')
    check_error = tree.xpath('//div[@class="error middle"]')
    for word in reseller_dict:
        if word in ads_title.lower():
            return ()
    if check_offers and int(check_offers[0]) > offers_count or check_error:
        return ()
    ads_price = tree.xpath(
        '//div[@class="price-seller"]/'
        '*[self::span[@class="price"] or self::span[@class="negotiated"]]'
        '/text()'
    )[0]
    ads_usd_price = 'Договорная'
    ads_eur_price = 'Договорная'
    ads_uah_price = 'Договорная'
    if ads_price != 'Договорная':
        ads_price_check = tree.xpath(
            '//div[@class="price-at-rate"]/span/@data-currency'
        )
        if 'USD' in ads_price_check and 'EUR' in ads_price_check:
            ads_usd_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="USD"]/'
                'text()'
            )[0].replace(' ', '')
            ads_eur_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="EUR"]/'
                'text()'
            )[0].replace(' ', '')
            ads_uah_price = ads_price.replace('грн', '').replace(' ', '')

        elif 'USD' in ads_price_check and 'UAH' in ads_price_check:
            ads_usd_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="USD"]/'
                'text()'
            )[0].replace(' ', '')
            ads_eur_price = ads_price.replace('€', '').replace(' ', '')
            ads_uah_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="UAH"]/'
                'text()'
            )[0].replace(' ', '')
        elif 'EUR' in ads_price_check and 'UAH' in ads_price_check:
            ads_usd_price = ads_price.replace('$', '').replace(' ', '')
            ads_eur_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="EUR"]/'
                'text()'
            )[0].replace(' ', '')
            ads_uah_price = tree.xpath(
                '//div[@class="price-at-rate"]/span[@data-currency="UAH"]/'
                'text()'
            )[0].replace(' ', '')
    try:
        ads_location = tree.xpath('//div[@class="location"]/text()')
        ads_location = ads_location[0].strip()
    except IndexError:
        ads_location = tree.xpath(
            '//div[@id="breadcrumbs"]/div[3]//span[@itemprop="title"]/text()'
        )[0].strip()
    ads_phone = tree.xpath(
        '//div[@class="phone conversion_phone_used"]//@data-phone-number'
    )
    ads_phone = [phone.replace(' ', '') for phone in ads_phone]
    ads_phone = ', '.join(ads_phone)
    ads_create = tree.xpath('//div[@id="addDate"]/span/text()')[0]
    return (
        ads_title,
        url,
        ads_usd_price,
        ads_eur_price,
        ads_uah_price,
        ads_location,
        ads_phone,
        ads_create
    )


def generate_pages_url(url, count):
    pages_list = []
    for page in range(count):
        page_url = url.replace('page=0', 'page={}'.format(page))
        pages_list.append(page_url)
    return pages_list


def request_thread_worker():
    while True:
        what_do, index, url = request_queue.get()
        if what_do == 'get_ads_url':
            url_list = get_ads_url(url)
            chunk_ads_url_list[index] = url_list
        elif what_do == 'get_ads_info':
            url_list = get_ads_info(url)
            chunk_ads_url_list[index] = url_list
        pbar.update(1)
        request_queue.task_done()


def start_thread_worker(worker_thread_count):
    for i in range(worker_thread_count):
        thread_worker = Thread(target=request_thread_worker)
        thread_worker.daemon = True
        thread_worker.start()


def url_list_to_thread(what_do, url_list):
    for index in range(len(url_list)):
        request_queue.put([what_do, index, url_list[index]])
    request_queue.join()


def chunkify(items, size):
    for i in range(0, len(items), size):
        yield items[i:i+size]


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Авто'
    head_table = [
        'Название',
        'Ссылка',
        'Цена в долларах',
        'Цена в евро',
        'Цена в гривнах',
        'Место',
        'Телефон',
        'Создан'
    ]
    ws.append(head_table)
    wb.save('auto.xlsx')


def fill_workbook(ads_info):
    wb = load_workbook('auto.xlsx')
    ws = wb.active
    for info in ads_info:
        if info:
            ws.append(info)
    wb.save('auto.xlsx')


def check_table_exist(table_list):
    for table in table_list:
        c.execute(
            'SELECT name FROM sqlite_master WHERE type="table" '
            'AND name="{}"'.format(table)
        )
        table_count = c.fetchall()
        if not table_count:
            if table == 'got_ads_url':
                c.execute('''CREATE TABLE got_ads_url
                             (url text PRIMARY KEY)''')
            if table == 'group_ads_phone':
                c.execute('''CREATE TABLE group_ads_phone
                                             (ads_phone text PRIMARY KEY,
                                              count text)''')
            if table == 'ads_info':
                c.execute('''CREATE TABLE ads_info
                             (title text,
                              url text PRIMARY KEY,
                              ads_usd_price text,
                              ads_eur_price text,
                              ads_uah_price text,
                              ads_location text,
                              ads_phone text,
                              ads_create text)''')
            if table == 'got_ads_info':
                c.execute('''CREATE TABLE got_ads_info
                             (title text,
                              url text PRIMARY KEY,
                              ads_usd_price text,
                              ads_eur_price text,
                              ads_uah_price text,
                              ads_location text,
                              ads_phone text,
                              ads_create text)''')


def get_improved_url(url):
    driver.get(url)
    wait.until(ec.element_to_be_clickable((
        By.CSS_SELECTOR, '#sortFilter > span'
    )))
    driver.find_element_by_xpath(
        '//*[@id="sortFilter"]/span[1]/a'
    ).click()
    wait.until(ec.element_to_be_clickable((
        By.CSS_SELECTOR, '#sortFilter span[class*="options"]'
    )))
    driver.find_element_by_xpath(
        '//*[@id="sortFilter"]/span[1]/span/a[1]'
    ).click()
    url = driver.current_url
    print(url)
    return url


if __name__ == '__main__':
    thread_count = 10       # Количество потоков
    offers_count = 1        # Количество объявлений продавца
    double_count = 1        # Количество объявлений с одинаковыми номерами
    chunk_size = 100        # Количество объявлений на которое разбиваеться
    reseller_dict = [       # Словарь фильтрации объявлений
        'official',
        'europe',
        'europa',
        'individual',
        'maximal',
        'maksimal',
        'top',
        'navi',
        'korea',
        'koreya',
        'exclusive',
        'premium',
        'panorama',
        'comfort',
        'executive',
        'special',
        'edition',
        'titanium',
        'ne krashen',
        'novaya',
        'gaz',
        '100%',
        '1000%',
        'original',
        'lux',
        'luxury',
        'obslugen',
        'obslugena',
        'ideal',
        'srochno',
        'gas',
        'full',
        'komfort',
        'avtomat',
        'torg',
        'официальная'
    ]
    request_queue = Queue()
    start_thread_worker(thread_count)
    conn = sqlite3.connect('auto.db')
    c = conn.cursor()
    check_table_exist(
        [
            'got_ads_url',
            'group_ads_phone',
            'got_ads_info',
            'ads_info'
        ]
    )
    c.execute('DELETE FROM got_ads_url')
    c.execute('DELETE FROM got_ads_info')
    c.execute('DELETE FROM group_ads_phone')
    conn.commit()
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(chrome_options=chrome_options)
    wait = WebDriverWait(driver, 15)
    start_url = input('Введите URL для поиска: ')
    start_time = time.time()
    if '&page=' not in start_url:
        start_url += '&page=0'
    else:
        start_url = re.sub(r'&size=\d*', '&size=0', start_url)
    print('Получаем количество страниц')
    page_count = get_page_count(start_url)
    print('Получаем исправленный URL')
    start_url = get_improved_url(start_url)
    driver.quit()
    page_list = generate_pages_url(start_url, page_count)
    pbar = tqdm.tqdm(total=len(page_list))
    pbar.set_description('Сбор URL объявлений со страниц')
    for chunk in chunkify(page_list, chunk_size):
        chunk_ads_url_list = [[] for x in chunk]
        url_list_to_thread('get_ads_url', chunk)

        chunk_ads_url_list = [
            (item,) for subdict in chunk_ads_url_list for item in subdict
        ]
        c.executemany(
            'INSERT OR IGNORE INTO got_ads_url VALUES (?)',
            chunk_ads_url_list
        )
        conn.commit()
    pbar.close()
    c.execute(
        'SELECT * FROM got_ads_url'
    )
    table_from_sql = c.fetchall()
    ads_url_list = [item for subdict in table_from_sql for item in subdict]
    pbar = tqdm.tqdm(total=len(ads_url_list))
    pbar.set_description('Сбор информации из объявлений')
    for chunk in chunkify(ads_url_list, chunk_size):
        chunk_ads_url_list = [[] for x in chunk]
        url_list_to_thread('get_ads_info', chunk)
        white_url_list = [tup for tup in chunk_ads_url_list if tup]
        c.executemany(
            'INSERT OR IGNORE INTO got_ads_info VALUES (?,?,?,?,?,?,?,?)',
            white_url_list
        )
        c.executemany(
            'INSERT OR IGNORE INTO ads_info VALUES (?,?,?,?,?,?,?,?)',
            white_url_list
        )
        conn.commit()
    pbar.close()
    c.execute(
        'SELECT ads_phone, COUNT(*) c FROM ads_info GROUP BY ads_phone '
        'HAVING c > {}'.format(double_count)
    )
    double_list = c.fetchall()
    c.executemany(
        'INSERT OR IGNORE INTO group_ads_phone VALUES (?,?)',
        double_list
    )
    conn.commit()
    c.execute(
        'SELECT * FROM got_ads_info WHERE NOT EXISTS '
        '(SELECT ads_phone FROM group_ads_phone '
        'WHERE got_ads_info.ads_phone = ads_phone) '
        'ORDER BY got_ads_info.title'
    )
    ads_list_from_sql = c.fetchall()
    create_workbook()
    fill_workbook(ads_list_from_sql)
    print(
        'Всего затрачено времени на сбор данных: {} минут'.format(
            round((time.time() - start_time)/60)
        )
    )
