import re
import pathlib
import pandas
import datetime
import requests
import user_agent
from tqdm import tqdm
from PIL import Image
from lxml import html
from io import BytesIO
from openpyxl import Workbook, load_workbook
from queue import Queue
from threading import Thread


def get_request(url):
    desktop_agent = user_agent.generate_user_agent(os=('win', 'mac', 'linux'))
    user_agent_header = {'User-Agent': desktop_agent}
    header = {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/67.0.3396.99 Mobile Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,'
                  'application/xml;q=0.9,image/webp,'
                  'image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'ru, en; q =0.9',
        'Connection': 'keep-alive'
    }
    header.update(user_agent_header)
    response = requests.get(url, headers=header, timeout=30)
    if response.status_code == requests.codes.ok:
        return response
    else:
        return None


def crate_image_dir(dir_name):
    pathlib.Path(dir_name).mkdir(parents=True, exist_ok=True)


def get_cryptocurrencies_url(url):
    response = get_request(url + '/all/views/all/')
    tree = html.fromstring(response.content)
    all_a_tag = tree.xpath('//tr/td[@class="no-wrap currency-name"]/a')
    return [start_url + a_tag.get('href') for a_tag in all_a_tag]


def get_ico(tree):
    ico_tag = tree.xpath('//h1/img')
    if ico_tag:
        ico = ico_tag[0].get('src')
        big_ico = ico.replace('32x32', '128x128')
    else:
        big_ico = None
    return big_ico


def get_name(tree):
    img_tag = tree.xpath('//h1[@class="details-panel-item--name"]/img')
    if img_tag:
        img_tag = img_tag[0].get('alt')
    else:
        img_tag = None
    return img_tag


def get_short_name(tree):
    span_tag = tree.xpath('//h1[@class="details-panel-item--name"]/span')
    if span_tag:
        span_tag = span_tag[0].text.replace('(', '').replace(')', '')
    else:
        span_tag = None
    return span_tag


def get_tag_value(tree, pattern):
    pattern_tag = tree.xpath(pattern)
    if pattern_tag:
        pattern_value = pattern_tag[0].text.strip()
    else:
        pattern_value = None
    return pattern_value


def get_details_panel_item_links(tree):
    li_dict = {}
    li_tags = tree.xpath(
        '//ul[contains(concat(" ",normalize-space(@class)," "),'
        '" details-panel-item--links ")]/li'
    )
    for li_tag in li_tags:
        tags = li_tag.xpath('child::span')
        if len(tags) > 1:
            tag_name = tags[0].get('title')
            tag_value = []
            for index in range(1, len(tags)):
                tag_value.append(tags[index].text.strip())
                li_dict.update({tag_name: ', '.join(tag_value)})
        else:
            tag_value = li_tag.xpath('child::a')
            li_dict.update({tag_value[0].text: tag_value[0].get('href')})
    return li_dict


def get_markets(tree):
    table_tag = tree.xpath('//table[@id="markets-table"]')
    if table_tag:
        dataframe = pandas.read_html(html.tostring(table_tag[0]))
    else:
        dataframe = None
    table_list = dataframe[0].to_dict(orient='records')
    table_string = [
        [
            str(row.get('#')),
            str(row.get('Source')),
            str(row.get('Pair')),
            str(row.get(
                'Volume (24h)'
            )).replace('*', '').replace('$', '').strip(),
            str(row.get('Price')).replace('*', '').replace('$', '').strip(),
            str(row.get('Volume (%)')).replace('%', '').strip(),
            str(row.get('Category')),
            str(row.get('Fee Type')),
            str(row.get('Updated'))
        ] for row in table_list
    ]
    return '\r\n'.join([';'.join(row) for row in table_string])


def get_historical_data(url):
    response = get_request(url + 'historical-data/')
    tree = html.fromstring(response.content)
    table_tag = tree.xpath('//div[@class="table-responsive"]/table')
    if table_tag:
        dataframe = pandas.read_html(html.tostring(table_tag[0]))
    else:
        dataframe = None
    table_list = dataframe[0].to_dict(orient='records')
    table_string = [
        [
            str(index),
            str(row.get('Date')),
            str(row.get('Open*')),
            str(row.get('High')),
            str(row.get('Low')),
            str(row.get('Close**')),
            str(row.get('Volume')),
            str(row.get('Market Cap'))
        ] for index, row in enumerate(table_list)
    ]
    return '\r\n'.join([';'.join(row) for row in table_string])


def get_social_twitter(tree):
    pattern_tag = tree.xpath('//a[@class="twitter-timeline"]')
    if pattern_tag:
        pattern_value = pattern_tag[0].get('href')
    else:
        pattern_value = None
    return pattern_value


def get_social_reddit(text):
    found = re.findall(r'oScript\.src = ["](.*)["]', text)
    if found:
        found = found[0].replace('.embed?limit=9', '')
    else:
        found = None
    return found


def get_charts(name):
    charts_list = []
    time_now = datetime.datetime.now()
    end_time = datetime.datetime.timestamp(time_now)
    time_delta = datetime.timedelta(days=365)
    start_time = end_time - time_delta.total_seconds()
    start_time = str(int(start_time)) + '000'
    end_time = str(int(end_time)) + '000'
    url = (
        'https://graphs2.coinmarketcap.com/currencies/'
        '{name}/{start_time}/{end_time}/'.format(
            name=name,
            start_time=start_time,
            end_time=end_time
        )
    )
    response = get_request(url)
    json_data = response.json()
    time_list = json_data['market_cap_by_available_supply']
    time_list = [current_time[0] for current_time in time_list]
    market_cap = {
        current_time[0]: current_time[1]
        for current_time
        in json_data['market_cap_by_available_supply']
    }
    price_btc = {
        current_time[0]: current_time[1]
        for current_time
        in json_data['price_btc']
    }
    price_usd = {
        current_time[0]: current_time[1]
        for current_time
        in json_data['price_usd']
    }
    volume_usd = {
        current_time[0]: current_time[1]
        for current_time
        in json_data['volume_usd']
    }
    for index, current_time in enumerate(time_list):
        charts_list.append(
            str(index+1) + ';' +
            str(datetime.datetime.fromtimestamp(
                current_time/1000
            ).strftime('%A;%b %d %Y;%H:%M:%S')) + ';' +
            str(market_cap[current_time]) + ';' +
            str(price_usd[current_time]) + ';' +
            str(price_btc[current_time]) + ';' +
            str(volume_usd[current_time])
        )
    return '\r\n'.join(charts_list)


def improve_name(text):
    chars = '\/:*?"<>|+%!@'
    for char in chars:
        text = text.replace(char, '')
    return text


def download_icon(url, name):
    extension = '.' + url.split('.')[-1].lower()
    image_name = '_'.join(name.lower().split())
    image_name = improve_name(image_name)
    save_path = 'logo/' + image_name + '_logo' + extension
    if not pathlib.Path(save_path).is_file():
        response = get_request(url)
        if response.status_code == requests.codes.ok:
            try:
                img = Image.open(BytesIO(response.content))
            except OSError:
                return None
            except OverflowError:
                return None
            except ValueError:
                return None
            try:
                img.save(save_path)
            except OSError:
                img = img.convert("RGB")
                img.save(save_path)
        else:
            save_path = None
    return save_path


def get_cryptocurrencie_info(url):
    response = get_request(url)
    tree = html.fromstring(response.content)
    ico_url = get_ico(tree)
    name = get_name(tree)
    short_name = get_short_name(tree)
    url_name = url.split('/')[-2]
    details_panel_item_links = get_details_panel_item_links(tree)
    current_price_usd = get_tag_value(tree, '//span[@id="quote_price"]/span')
    current_price_btc = get_tag_value(
        tree,
        '//div[contains(concat(" ",normalize-space(@class)," "),'
        '" details-panel-item--price ")]/span[@class="text-gray"][1]/span'
    )
    current_price_eth = get_tag_value(
        tree,
        '//div[contains(concat(" ",normalize-space(@class)," "),'
        '" details-panel-item--price ")]/span[@class="text-gray"][2]/span'
    )
    market_cap_usd = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Market Cap"]/div/span[1]/span'
    )
    market_cap_btc = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Market Cap"]/div/span[2]/span'
    )
    market_cap_eth = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Market Cap"]/div/span[3]/span'
    )
    volume_24h_usd = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Volume (24h)"]/div/span[1]'
        '/span'
    )
    volume_24h_btc = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Volume (24h)"]/div/span[2]'
        '/span'
    )
    volume_24h_eth = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Volume (24h)"]/div/span[3]'
        '/span'
    )
    circulating_supply = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Circulating Supply"]/div/span'
    )
    total_supply = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Total Supply"]/div/span'
    )
    max_supply = get_tag_value(
        tree,
        '//div[@class="coin-summary-item"][h5 = "Max Supply"]/div/span'
    )
    charts = get_charts(url_name)
    markets = get_markets(tree)
    social_twitter = get_social_twitter(tree)
    social_reddit = get_social_reddit(response.text)
    historical_data = get_historical_data(url)
    ico_path = download_icon(ico_url, name)
    return {
        'ico': ico_path,
        'name': name,
        'short_name': short_name,
        'url_name': url_name,
        'details_panel_item_links': details_panel_item_links,
        'current_price_usd': current_price_usd,
        'current_price_btc': current_price_btc,
        'current_price_eth': current_price_eth,
        'market_cap_usd': market_cap_usd,
        'market_cap_btc': market_cap_btc,
        'market_cap_eth': market_cap_eth,
        'volume_24h_usd': volume_24h_usd,
        'volume_24h_btc': volume_24h_btc,
        'volume_24h_eth': volume_24h_eth,
        'circulating_supply': circulating_supply,
        'total_supply': total_supply,
        'max_supply': max_supply,
        'charts': charts,
        'markets': markets,
        'social_twitter': social_twitter,
        'social_reddit': social_reddit,
        'historical_data': historical_data
    }


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project'
    head_table = [
        'ICO Name',
        'Financial/Token info/Token',
        'Rank',
        'Website link',
        'Announcement',
        'Explorer link',
        'Explorer 2 link',
        'Explorer 3 link',
        'Message Board',
        'Message Board 2',
        'Chat',
        'Chat 2',
        'Source Code link',
        'Technical Documentation',
        'Tags',
        'Current price, $ (USD)',
        'Current price, BTC',
        'Current price, ETH',
        'Market Cap, USD',
        'Market Cap, BTC',
        'Market Cap, ETH',
        'Volume (24h), USD',
        'Volume (24h), BTC',
        'Volume (24h), ETH',
        'Circulating Supply',
        'Total Supply',
        'Max Supply',
        'Chart',
        'Markets',
        'Social [twitter]',
        'Social [reddit]',
        'Historical data'
    ]
    ws.append(head_table)
    wb.save('project.xlsx')


def fill_workbook(projects):
    wb = load_workbook('project.xlsx')
    ws = wb.active
    for project in projects:
        ws.append([
            project['name'],
            project['short_name'],
            project['details_panel_item_links'].get('Rank'),
            project['details_panel_item_links'].get('Website'),
            project['details_panel_item_links'].get('Announcement'),
            project['details_panel_item_links'].get('Explorer'),
            project['details_panel_item_links'].get('Explorer 2'),
            project['details_panel_item_links'].get('Explorer 3'),
            project['details_panel_item_links'].get('Message Board'),
            project['details_panel_item_links'].get('Message Board 2'),
            project['details_panel_item_links'].get('Chat'),
            project['details_panel_item_links'].get('Chat 2'),
            project['details_panel_item_links'].get('Source Code'),
            project['details_panel_item_links'].get(
                'Technical Documentation'
            ),
            project['details_panel_item_links'].get('Tags'),
            project['current_price_usd'],
            project['current_price_btc'],
            project['current_price_eth'],
            project['market_cap_usd'],
            project['market_cap_btc'],
            project['market_cap_eth'],
            project['volume_24h_usd'],
            project['volume_24h_btc'],
            project['volume_24h_eth'],
            project['circulating_supply'],
            project['total_supply'],
            project['max_supply'],
            project['charts'],
            project['markets'],
            project['social_twitter'],
            project['social_reddit'],
            project['historical_data']
        ])
    wb.save('project.xlsx')


def thread_worker():
    while True:
        index, url = threads_queue.get()
        info = get_cryptocurrencie_info(url)
        temp_thread[index] = info
        pbar.update(1)
        threads_queue.task_done()


def start_thread_pool(worker_thread_count):
    for i in range(worker_thread_count):
        t = Thread(target=thread_worker)
        t.setDaemon(True)
        t.start()


def threading_pages_list(url_list):
    for index in range(len(url_list)):
        threads_queue.put([index, url_list[index]])
    threads_queue.join()


def chunkify(items, size):
    for i in range(0, len(items), size):
        yield items[i:i+size]


if __name__ == '__main__':
    threads_queue = Queue()
    thread_responses_list = []
    worker_count = 8
    chunk_size = 100
    start_thread_pool(worker_count)
    crate_image_dir('logo')
    start_url = 'https://coinmarketcap.com'
    cryptocurrencies_url = get_cryptocurrencies_url(start_url)
    pbar = tqdm(
        total=len(cryptocurrencies_url),
        desc='Сбор информации о проекте'
    )
    create_workbook()
    for chunk in chunkify(cryptocurrencies_url, chunk_size):
        temp_thread = [[] for x in chunk]
        threading_pages_list(chunk)
        fill_workbook(temp_thread)
    pbar.close()
